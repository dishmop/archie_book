#!/usr/bin/env python3
"""
Convert image_submission_final.csv to Excel with formatting:
- Bold headings
- Auto-fit column widths
- Conditional formatting for Red/Green/Amber copyright status
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
CSV_FILE = PROJECT_DIR / "image_submission.csv"
EXCEL_FILE = PROJECT_DIR / "image_submission.xlsx"

# Color fills for copyright status
FILLS = {
    "Red": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "Green": PatternFill(start_color="7BC96F", end_color="7BC96F", fill_type="solid"),
    "Amber": PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid"),
}

def main():
    # Read CSV
    with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print("CSV is empty")
        return

    headers = rows[0]
    data = rows[1:]

    # Find copyright_status column index
    try:
        status_col = headers.index("copyright_status") + 1  # 1-indexed for openpyxl
    except ValueError:
        status_col = None

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Submission"

    # Track max width for each column
    col_widths = [0] * len(headers)

    # Write headers (bold)
    bold_font = Font(bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold_font
        col_widths[col - 1] = len(str(header))

    # Find numeric columns
    try:
        chapter_col = headers.index("chapter") + 1
    except ValueError:
        chapter_col = None
    try:
        figure_col = headers.index("figure") + 1
    except ValueError:
        figure_col = None

    # Write data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            # Convert chapter and figure columns to integer when purely numeric
            if col_idx in (chapter_col, figure_col) and value.isdigit():
                value = int(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Update max width
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))

            # Apply color fill for copyright_status column
            if col_idx == status_col and value in FILLS:
                cell.fill = FILLS[value]

            # Left-align figure column
            if col_idx == figure_col:
                cell.alignment = Alignment(horizontal="left")

    # Set column widths (add padding)
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 80)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Count status values
    status_counts = {"Green": 0, "Amber": 0, "Red": 0}
    if status_col:
        for row in data:
            if len(row) >= status_col:
                status = row[status_col - 1]  # 0-indexed for list
                if status in status_counts:
                    status_counts[status] += 1

    # Create summary sheet
    summary = wb.create_sheet(title="Summary")

    # Header
    summary.cell(row=1, column=1, value="Status").font = bold_font
    summary.cell(row=1, column=2, value="Count").font = bold_font

    # Data rows with colors
    for i, (status, count) in enumerate(status_counts.items(), 2):
        cell_status = summary.cell(row=i, column=1, value=status)
        cell_status.fill = FILLS[status]
        summary.cell(row=i, column=2, value=count)

    # Total row
    total_row = len(status_counts) + 2
    summary.cell(row=total_row, column=1, value="Total").font = bold_font
    summary.cell(row=total_row, column=2, value=sum(status_counts.values())).font = bold_font

    # Set column widths
    summary.column_dimensions["A"].width = 12
    summary.column_dimensions["B"].width = 10

    # Save
    wb.save(EXCEL_FILE)
    print(f"Created: {EXCEL_FILE}")
    print(f"Summary: Green={status_counts['Green']}, Amber={status_counts['Amber']}, Red={status_counts['Red']}")

if __name__ == "__main__":
    main()
